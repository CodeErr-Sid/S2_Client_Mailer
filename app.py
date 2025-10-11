# app_final.py
import streamlit as st
import json
import pandas as pd
import os
from pathlib import Path
import hashlib
from datetime import datetime
import smtplib
from email.message import EmailMessage
import requests
from openpyxl import load_workbook

# --- CONFIG ---
st.set_page_config(page_title="S2 Client Recievable's", page_icon=r"assets\s2logo.png", layout="wide")

# --- STYLE ---
st.markdown("""
    <style>
    body {
        background-color: #0d1117;
        color: white;
        font-family: 'Montserrat', sans-serif;
    }
    .stApp {
        background-color: #0d1117;
    }
    .login-logo {
        width: 90px;
        height: auto;
        margin-bottom: 1rem;
        border-radius: 50%;
        box-shadow: 0 0 15px rgba(0, 193, 110, 0.3);
    }
    .login-title {
        font-size: 1.8rem;
        font-weight: 600;
        margin-bottom: 1rem;
        color: #f1f1f1;
    }
    .stTextInput>div>div>input {
        background-color: #0d1117;
        color: #f1f1f1;
        border: 1px solid #30363d;
        border-radius: 6px;
    }

    </style>
""", unsafe_allow_html=True)

# --- PASSWORD HASH ---
HASHED_PASSWORD = "1d493066e5f3f142eb6c9efae9511745afbc03286e64ae192c3ef0b420cd9019"

def check_password(input_password):
    hashed_input = hashlib.sha256(input_password.encode()).hexdigest()
    return hashed_input == HASHED_PASSWORD

# --- LOGO PATH ---
logo_path = Path(__file__).parent / "assets" / "s2logo.png"

# --- LIVE RATE + CURRENCY HELPERS ---
def get_live_usd_to_inr_rate(default_rate=83.0):
    """Fetch live USD->INR rate (exchangerate.host). Fallback to default_rate."""
    try:
        url = "https://api.exchangerate.host/latest?base=USD&symbols=INR"
        r = requests.get(url, timeout=5)
        data = r.json()
        if isinstance(data, dict) and "rates" in data and "INR" in data["rates"]:
            return float(data["rates"]["INR"])
    except Exception:
        pass
    return default_rate

if "USD_TO_INR" not in st.session_state:
    st.session_state.USD_TO_INR = get_live_usd_to_inr_rate()

def parse_currency_from_string(s):
    """Parse currency and numeric amount from a string like '$1,200' or '₹4,294.00'.
    Returns (currency, amount) where currency is 'USD' or 'INR'."""
    if s is None:
        return "INR", 0.0
    if isinstance(s, (int, float)):
        # numeric with no symbol — assume INR
        return "INR", float(s)
    text = str(s).strip()
    # If the string already starts with symbol
    if text.startswith("$"):
        try:
            amt = float(text.replace("$", "").replace(",", "").strip())
        except:
            amt = 0.0
        return "USD", amt
    if text.startswith("₹") or text.lower().startswith("rs"):
        try:
            amt = float(text.replace("₹", "").replace("Rs", "").replace("rs", "").replace(",", "").strip())
        except:
            amt = 0.0
        return "INR", amt
    # fallback: try to parse numbers (assume INR)
    try:
        return "INR", float(text.replace(",", ""))
    except:
        return "INR", 0.0

def parse_currency(value):
    """Generic parse function (accepts numeric or string)."""
    return parse_currency_from_string(value)

def convert_to_inr(currency, amount):
    if currency == "USD":
        return amount * st.session_state.USD_TO_INR
    return amount

def format_symbol_amount(symbol, amount):
    return f"{symbol}{amount:,.2f}"

# --- EXCEL DISPLAY HELPER (uses openpyxl to preserve formatting) ---
def read_excel_with_display_values(path, sheet_name=None):
    """
    Read workbook using pandas (for data) and openpyxl to produce 'display' strings for currency columns.
    Returns (df, display_map) where display_map is dict: column_name -> list of display strings
    """
    # Read using pandas to preserve headers and data
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", dtype=object)

    # Use openpyxl to read raw cell formats and values
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Map header names to column letters
    headers = {}
    first_row = next(ws.iter_rows(min_row=1, max_row=1))
    for cell in first_row:
        if cell.value is not None:
            headers[cell.column_letter] = str(cell.value)

    # Build reverse map: header_name -> column_letter
    header_to_col = {v: k for k, v in headers.items()}

    display_map = {}
    # We'll try to create a display mapping only for numeric columns that might be currency.
    # For each column name present in df, find its column letter and get display strings.
    for col_name in df.columns:
        # find column letter for this header (exact match)
        col_letter = None
        for letter, h in headers.items():
            # header matching ignoring case/whitespace
            if str(h).strip().lower() == str(col_name).strip().lower():
                col_letter = letter
                break
        if not col_letter:
            # skip — we can't map to openpyxl column
            continue

        # iterate cells under this column (starting row 2)
        disp_values = []
        row_idx = 2
        for r in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            # find the cell in this row for our column
            # map column_letter to index: openpyxl column letters to index
            from openpyxl.utils import column_index_from_string
            col_idx = column_index_from_string(col_letter) - 1
            try:
                cell = r[col_idx]
            except Exception:
                disp_values.append(None)
                row_idx += 1
                continue

            # Determine display string:
            cval = cell.value
            nf = (cell.number_format or "").upper()
            # If cell has a number format containing $ or ₹ or "[$" patterns
            # common number_format examples: '"$"#,##0.00', '$#,##0.00', '₹#,##0.00', '₹ #,##0.00;[Red]-₹ #,##0.00'
            display = None
            try:
                if cval is None:
                    display = None
                else:
                    # If number format contains $, treat as USD
                    if "$" in nf or "[$USD" in nf or '"$"' in nf:
                        # format numeric with two decimals and thousands
                        try:
                            display = f"${float(cval):,.2f}"
                        except:
                            display = str(cval)
                    elif "₹" in nf or "RS" in nf or "INR" in nf or "[$INR" in nf or '"₹"' in nf:
                        try:
                            display = f"₹{float(cval):,.2f}"
                        except:
                            display = str(cval)
                    else:
                        # if the cell value is string already containing symbols, preserve it
                        if isinstance(cval, str) and (cval.strip().startswith("$") or cval.strip().startswith("₹")):
                            display = cval.strip()
                        else:
                            # fallback: if numeric, format without symbol (we won't invent symbol for display_map)
                            if isinstance(cval, (int, float)):
                                display = f"{float(cval):,.2f}"
                            else:
                                display = str(cval)
            except Exception:
                display = str(cval)
            disp_values.append(display)
            row_idx += 1

        display_map[col_name] = disp_values

    wb.close()
    return df, display_map

# --- LOGIN LOGIC ---
if "logged_in" not in st.session_state:
    with st.container():
        st.markdown('<div class="login-container">', unsafe_allow_html=True)

        # ✅ Display Local Image (removed deprecated param)
        try:
            st.image(str(logo_path), width=90)
        except Exception:
            # ignore missing image
            pass

        st.markdown('<div class="login-title">🔐 Login to Continue!</div>', unsafe_allow_html=True)
        password = st.text_input("Enter Password", type="password", label_visibility="collapsed")
        st.write("")  # spacing

        if st.button("Login"):
            if check_password(password):
                st.session_state.logged_in = True
                st.rerun()
            else:
                st.error("❌ Incorrect password. Please try again.")
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# --- Page Config ---
# st.set_page_config(page_title="Invoice Tracker", layout="wide")
st.title("📜 S2 Inv Receivable's")

# --- Folders and Files ---
DATA_FOLDER = "data"
os.makedirs(DATA_FOLDER, exist_ok=True)
DATA_FILE = os.path.join(DATA_FOLDER, "last_uploaded.xlsx")
TIME_FILE = os.path.join(DATA_FOLDER, "upload_time.txt")
CREDENTIALS_FILE = os.path.join(DATA_FOLDER, "sender_credentials.txt")
CLIENT_EMAIL_FILE = os.path.join(DATA_FOLDER, "client_emails.json")

# (keep st.secrets usage)
sender_email = st.secrets.get("sender_email", "")
sender_password = st.secrets.get("sender_password", "")

# --- Session State Defaults ---
if "sender_email" not in st.session_state:
    st.session_state.sender_email = ""
if "sender_password" not in st.session_state:
    st.session_state.sender_password = ""
if "show_sender_modal" not in st.session_state:
    st.session_state.show_sender_modal = False
if "approver_action" not in st.session_state:
    st.session_state.approver_action = None

# --- Load Sender Credentials if exist ---
if os.path.exists(CREDENTIALS_FILE):
    with open(CREDENTIALS_FILE, "r") as f:
        lines = f.read().splitlines()
        if len(lines) >= 2:
            st.session_state.sender_email = lines[0]
            st.session_state.sender_password = lines[1]

# --- Top-right sender credentials modal toggle ---
if "show_sender_modal" not in st.session_state:
    st.session_state.show_sender_modal = False
if "sender_email_temp" not in st.session_state:
    st.session_state.sender_email_temp = st.session_state.get("sender_email", "")
if "sender_password_temp" not in st.session_state:
    st.session_state.sender_password_temp = st.session_state.get("sender_password", "")

# --- Trigger the modal toggle button ---
if st.session_state.show_sender_modal:
    with st.container():
        st.markdown("### ✉️ Configure Sender Credentials")

        # These fields persist across reruns
        st.session_state.sender_email_temp = st.text_input(
            "Sender Email",
            value=st.session_state.sender_email_temp,
            placeholder="example@gmail.com",
            key="sender_email_input"
        )

        st.session_state.sender_password_temp = st.text_input(
            "Password / App Password",
            value=st.session_state.sender_password_temp,
            type="password",
            placeholder="Enter your Gmail App Password",
            key="sender_password_input"
        )

        col_save, col_close = st.columns(2)

        with col_save:
            if st.button("💾 Save Credentials", key="save_credentials_btn"):
                st.session_state.sender_email = st.session_state.sender_email_temp.strip()
                st.session_state.sender_password = st.session_state.sender_password_temp.strip()

                # Save to file
                try:
                    with open(CREDENTIALS_FILE, "w") as f:
                        f.write(f"{st.session_state.sender_email}\n{st.session_state.sender_password}")
                    success_box = st.empty()
                    # success_box.success("")
                    st.toast(f"Sender credentials saved successfully!", icon="✅")
                    import time
                    time.sleep(1.5)
                    success_box.empty()
                except Exception as e:
                    st.error(f"❌ Failed to save credentials: {e}")

        with col_close:
            if st.button("❌ Close", key="close_credentials_btn"):
                st.session_state.show_sender_modal = False
else:
    # Show button to open modal (outside the modal logic)
    if st.button("🔑 Modify Sender"):
        st.session_state.show_sender_modal = True
        st.rerun()



# --- Load Client Emails safely ---
if "client_emails" not in st.session_state or not isinstance(st.session_state.client_emails, dict):
    st.session_state.client_emails = {"cc_email": "", "clients": {}}
else:
    st.session_state.client_emails.setdefault("cc_email", "")
    st.session_state.client_emails.setdefault("clients", {})

if os.path.exists(CLIENT_EMAIL_FILE):
    try:
        with open(CLIENT_EMAIL_FILE, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if content:
                try:
                    loaded = json.loads(content)
                    if isinstance(loaded, dict):
                        loaded.setdefault("cc_email", "")
                        loaded.setdefault("clients", {})
                        st.session_state.client_emails = loaded
                except json.JSONDecodeError:
                    st.session_state.client_emails = {"cc_email": "", "clients": {}}
            else:
                st.session_state.client_emails = {"cc_email": "", "clients": {}}
    except Exception:
        st.session_state.client_emails = {"cc_email": "", "clients": {}}

# --- Top-right buttons for Sender + Client Email ---
col1, col2 = st.columns([0.3, 0.3])

# with col1:
#     if st.button("✉︎ Configure Sender Credentials", key="sender_modal_btn"):
#         st.session_state.show_sender_modal = True
#         st.rerun()

with col1:
    if st.button("🗝️ Modify Clients", key="client_modal_btn"):
        st.session_state.show_client_email_modal = True
        st.rerun()

# --- Client Email Configuration Modal ---
if st.session_state.get("show_client_email_modal", False):
    with st.container():
        st.markdown("""<div>""", unsafe_allow_html=True)
        st.markdown("### 📧 Client Email Configurator")
        st.markdown("Add or edit client email IDs here once — they’ll auto-fill for emails.")

        cc_mail_input = st.text_input("Global CC Email", value=st.session_state.client_emails.get("cc_email", ""), key="cc_email_input")

        if st.session_state.get("stored_data", None) is not None:
            df_clients = st.session_state.stored_data
            client_col = next((col for col in df_clients.columns if "client name" in col.lower()), None)
            if client_col:
                unique_clients = sorted(df_clients[client_col].dropna().unique().tolist())
                st.markdown("#### ✉️ Clients & Emails")
                st.session_state.client_emails.setdefault("clients", {})
                for client in unique_clients:
                    prev_email = st.session_state.client_emails["clients"].get(client, "")
                    new_val = st.text_input(f"{client}", value=prev_email, key=f"client_email_{client}")
                    st.session_state.client_emails["clients"][client] = new_val
            else:
                st.warning("⚠️ No 'Client Name' column found in Excel.")
        else:
            st.info("📂 Upload Excel first to load clients.")

        col_save, col_close = st.columns(2)
        with col_save:
            if st.button("💾 Save All", key="save_client_emails"):
                st.session_state.client_emails["cc_email"] = cc_mail_input or ""
                st.session_state.client_emails.setdefault("clients", {})
                try:
                    with open(CLIENT_EMAIL_FILE, "w", encoding="utf-8") as f:
                        json.dump(st.session_state.client_emails, f, indent=4, ensure_ascii=False)
                    # st.success("")
                    st.toast(f"Client emails saved successfully!", icon="✅")

                except Exception as e:
                    st.error(f"❌ Failed to save client emails: {e}")
        with col_close:
            if st.button("❌ Close", key="close_client_email_modal"):
                st.session_state.show_client_email_modal = False

        st.markdown("</div>", unsafe_allow_html=True)

# --- Email Sending Function ---
def send_email(sender_email, sender_password, to_email, subject, body, cc=None):
    msg = EmailMessage()
    msg['Subject'] = subject
    msg['From'] = sender_email
    msg['To'] = to_email
    if cc:
        msg['Cc'] = cc
    msg.set_content(body)
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(sender_email, sender_password)
            smtp.send_message(msg)
        return True, "Email sent successfully!"
    except Exception as e:
        return False, str(e)

# --- Load Excel Data & Preserve Display Values ---
# st.sidebar.markdown("## ⚙️ Options")
# --- Manual USD→INR Exchange Rate Setting ---
RATE_FILE = os.path.join(DATA_FOLDER, "usd_inr_rate.txt")

# Load saved rate from file (if exists)
if os.path.exists(RATE_FILE):
    try:
        with open(RATE_FILE, "r") as f:
            saved_rate = float(f.read().strip())
            st.session_state.USD_TO_INR = saved_rate
    except Exception:
        pass  # fallback to session/default rate

# Input field to set and save custom rate
st.sidebar.markdown("### 💱 USD → INR Conversion Rate")
manual_rate = st.sidebar.number_input(
    "Set Exchange Rate (₹ per $)",
    min_value=50.0,
    max_value=150.0,
    value=float(st.session_state.get("USD_TO_INR", 83.0)),
    step=0.1,
)

# Save rate button
if st.sidebar.button("✅ Save"):
    st.session_state.USD_TO_INR = manual_rate
    with open(RATE_FILE, "w") as f:
        f.write(str(manual_rate))
    st.toast(f"✅ USD → INR rate saved: ₹{manual_rate:.2f}", icon="💾")

# Display current rate
# st.sidebar.info(f"Current USD → INR rate: ₹{st.session_state.USD_TO_INR:.2f}")


if os.path.exists(DATA_FILE):
    # read using pandas (object dtype) to avoid automatic cast where possible
    try:
        # read with openpyxl engine to be consistent
        df_raw = pd.read_excel(DATA_FILE, engine="openpyxl", dtype=object)
    except Exception:
        df_raw = pd.read_excel(DATA_FILE, dtype=object)
    # read display map using openpyxl
    try:
        df_loaded, display_map = read_excel_with_display_values(DATA_FILE)
    except Exception:
        df_loaded = df_raw.copy()
        display_map = {}
    st.session_state.stored_data = df_loaded
    if os.path.exists(TIME_FILE):
        with open(TIME_FILE, "r") as f:
            st.session_state.last_uploaded_time = f.read().strip()
else:
    st.session_state.stored_data = None
    st.session_state.last_uploaded_time = None

# --- File Upload Section ---
uploaded_file = st.file_uploader("Upload Excel File", type=["xlsx"])
if uploaded_file is not None:
    # Save uploaded file
    upload_path = os.path.join(DATA_FOLDER, "last_uploaded.xlsx")
    with open(upload_path, "wb") as f:
        f.write(uploaded_file.getbuffer())

    # Load the data
    try:
        df_raw = pd.read_excel(upload_path, engine="openpyxl", dtype=object)
    except Exception:
        df_raw = pd.read_excel(upload_path, dtype=object)

    # Read display map
    try:
        df_loaded, display_map = read_excel_with_display_values(upload_path)
    except Exception:
        df_loaded = df_raw.copy()
        display_map = {}

    # Save to DATA_FILE
    df_loaded.to_excel(DATA_FILE, index=False)
    st.session_state.stored_data = df_loaded

    # Save current upload time
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(TIME_FILE, "w") as f:
        f.write(current_time)
    st.session_state.last_uploaded_time = current_time

    st.toast(f"✅ File Uploaded Successfully!", icon="💾")

# --- Display Last Uploaded / Updated Time ---
if st.session_state.last_uploaded_time:
    st.info(f"📅 Last Updated : {st.session_state.last_uploaded_time}")


# --- Main Dashboard ---
if st.session_state.stored_data is not None:
    df = st.session_state.stored_data.copy()

    # Columns detection (same as before)
    paid_col = next((col for col in df.columns if "paid" in col.lower()), None)
    due_col = next((col for col in df.columns if "due" in col.lower()), None)
    amount_col = next((col for col in df.columns if any(x in col.lower() for x in ["amount", "total", "value"])), None)
    date_col = next((col for col in df.columns if "date" in col.lower() or "raised" in col.lower()), None)
    client_col = next((col for col in df.columns if "client name" in col.lower()), None)
    approver_mail_col = next((col for col in df.columns if "approver mail" in col.lower()), None)
    client_mail_col = next((col for col in df.columns if "client mail" in col.lower()), None)
    cc_mail_col = next((col for col in df.columns if "cc" in col.lower()), None)
    invoice_col = next((col for col in df.columns if "invoice" in col.lower() and ("no" in col.lower() or "number" in col.lower() or "id" in col.lower())), None)

    # Filter by client
    client_options = ["All Clients"] + sorted(df[client_col].dropna().unique().tolist()) if client_col else ["All Clients"]
    selected_client = st.selectbox("Select Client", client_options)
    df_filtered = df[df[client_col] == selected_client] if selected_client != "All Clients" else df.copy()

# --- Metrics (paid/unpaid detection stays the same with numeric fallback) ---
if paid_col:
    df_filtered[paid_col] = pd.to_numeric(df_filtered[paid_col], errors="coerce").fillna(0)
    paid_df = df_filtered[df_filtered[paid_col] > 0]
    unpaid_df = df_filtered[df_filtered[paid_col] == 0]
elif "due" in df.columns.str.lower().tolist():
    due_col = next((col for col in df.columns if "due" in col.lower()), None)
    df_filtered[due_col] = pd.to_numeric(df_filtered[due_col], errors="coerce").fillna(0)
    paid_df = df_filtered[df_filtered[due_col] == 0]
    unpaid_df = df_filtered[df_filtered[due_col] > 0]
else:
    paid_df = pd.DataFrame()
    unpaid_df = df_filtered

# --- Dashboard Summary ---
st.markdown("## ⌨ Dashboard")
col1, col2, col3, col4, col5, = st.columns(5)
col1.metric("🧾 Total Clients", df[client_col].nunique() if client_col else len(df))
col2.metric("📄 Total Invoices", len(df_filtered))
col3.metric("✅ Paid", len(paid_df))
col4.metric("⚠️ Pending", len(unpaid_df))

st.markdown(
    """
    <style>
    .st-emotion-cache-1q82h82 {
        overflow: visible !important;
        white-space: normal !important;
        text-overflow: clip !important;
        
    }
        .st-emotion-cache-efbu8t {
        background: linear-gradient(90deg, #ff5f6d, #ffc371, #24c6dc, #514a9d);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-weight: bold;       /* optional */
        font-size: 1.4rem;       /* adjust size if needed */
        # overflow: visible !important;
        # white-space: normal !important;
        # text-overflow: clip !important;
    }
    </style>
    """,
    unsafe_allow_html=True
)


# --- Total Due (converted to INR) ---
if due_col:
    unpaid_calc = unpaid_df.copy()

    # Ensure Currency column exists
    if "Currency" not in unpaid_calc.columns:
        unpaid_calc["Currency"] = unpaid_calc[amount_col].apply(lambda x: parse_currency(x)[0])

    # Convert each due amount properly based on currency
    def convert_due_to_inr(row):
        c, a = parse_currency(row[due_col])
        # If there's a separate Currency column, override detected currency
        if pd.notna(row.get("Currency")) and row["Currency"] in ["$", "USD", "usd"]:
            c = "USD"
        elif row.get("Currency") in ["₹", "INR", "inr"]:
            c = "INR"
        return convert_to_inr(c, a)

    unpaid_calc["INR_Value"] = unpaid_calc.apply(convert_due_to_inr, axis=1)

    total_due_inr = unpaid_calc["INR_Value"].sum()
    col5.metric("💰 Total Due (in ₹)", f"₹{total_due_inr:,.2f}")
else:
    col5.metric("💰 Total Due", "₹0.00")

import plotly.express as px

# --- Ageing Table & Graph ---
if unpaid_df.shape[0] > 0 and date_col:
    ageing_df = unpaid_df.copy()
    ageing_df[date_col] = pd.to_datetime(ageing_df[date_col], errors="coerce")
    ageing_df["Days Pending"] = (datetime.now() - ageing_df[date_col]).dt.days

    # Format invoice date column as DD-MMM-YYYY
    ageing_df[date_col] = ageing_df[date_col].dt.strftime("%d-%b-%Y")

    # Step 3: Ensure Currency column is included
    if "Currency" in df_filtered.columns:
        ageing_df = ageing_df.copy()
        ageing_df.loc[:, "Currency"] = df_filtered.loc[ageing_df.index, "Currency"]
    else:
        ageing_df.loc[:, "Currency"] = ageing_df[amount_col].apply(lambda x: parse_currency(x)[0])

    # Insert Currency Display before amount column
    if amount_col in ageing_df.columns:
        if amount_col in display_map:
            disp_vals = display_map[amount_col][:len(ageing_df)]
        else:
            disp_vals = ageing_df[amount_col].apply(lambda x: f"{x:,.2f}" if pd.notnull(x) else "")

        ageing_df.insert(
            ageing_df.columns.get_loc(amount_col),
            "Currency Display",
            ageing_df["Currency"]
        )

    # --- Display ---
    st.markdown("### ⊞ Ageing Table")
    st.dataframe(
        ageing_df[[client_col, invoice_col, "Currency", amount_col, date_col, "Days Pending"]],
        width="stretch"
    )

    st.markdown("### ☰ Ageing Graph")
    chart_df = ageing_df.copy()
    chart_df[invoice_col] = chart_df[invoice_col].astype(str)

    fig = px.bar(
        chart_df,
        x=invoice_col,
        y="Days Pending",
        text="Days Pending",
        title="Pending Days by Invoice",
    )

    fig.update_traces(textposition="outside", marker_color="#f7941d",)
    fig.update_layout(
        xaxis=dict(fixedrange=True, color="white", showgrid=True, gridcolor="rgba(255,255,255,0.2)"),
        yaxis=dict(fixedrange=True, color="white", showgrid=True, gridcolor="rgba(255,255,255,0.2)"),
        height=520,
        margin=dict(l=40, r=40, t=60, b=80),
        plot_bgcolor="black",
        paper_bgcolor="#0e1117",
        font=dict(color="white", size=14),
        title=dict(x=0.35, font=dict(size=20, color="#B2FFFF")),
    )
    st.plotly_chart(fig, config={"responsive": True}, key="ageing_chart")

else:
    st.info("✅ No pending invoices available — ageing analysis not applicable.")

# --- Pie Chart: Pending Invoices by Client ---
if unpaid_df.shape[0] > 0 and client_col:
    st.markdown("### ◔ Pending Invoices by Client")
    pending_by_client = unpaid_df.groupby(client_col).size().reset_index(name="Pending Count")
    fig_client_pie = px.pie(
        pending_by_client,
        names=client_col,
        values="Pending Count",
        title="Pending Invoices Distribution by Client",
        hole=0.4,
    )
    fig_client_pie.update_traces(textinfo="percent+label", textfont_size=14, marker=dict(line=dict(color='#0d1117', width=2)))
    fig_client_pie.update_layout(paper_bgcolor="#0e1117", plot_bgcolor="#0e1117", font=dict(color="white"), title=dict(x=0.35, font=dict(size=20, color="#B2FFFF")))
    st.plotly_chart(fig_client_pie, config={"responsive": True}, key="pending_client_chart")

# --- Pie Chart: Paid vs Pending ---
if len(paid_df) > 0 or len(unpaid_df) > 0:
    st.markdown("### ◔ Invoice Status Breakdown")
    pie_data = pd.DataFrame({"Status": ["Paid", "Pending"], "Count": [len(paid_df), len(unpaid_df)]})
    fig_pie = px.pie(pie_data, names="Status", values="Count", title="Paid vs Pending Invoices", hole=0.4, color="Status", color_discrete_map={"Paid": "#00C851", "Pending": "#FF4444"})
    fig_pie.update_traces(textinfo="percent+label", textfont_size=14)
    fig_pie.update_layout(paper_bgcolor="#0e1117", plot_bgcolor="#0e1117", font=dict(color="white"), title=dict(x=0.35, font=dict(size=20, color="#B2FFFF")))
    st.plotly_chart(fig_pie, config={"responsive": True}, key="status_chart")

# --- Email Actions Sidebar ---
if client_col and st.session_state.stored_data is not None:
    st.sidebar.markdown("### 📧 Email Actions")
    selected_client_name = st.sidebar.selectbox("Select Client for Email", sorted(df[client_col].dropna().unique().tolist()), key="client_selector")
    client_invoices = df[df[client_col] == selected_client_name].copy()

    # For due filter, prefer the due_col if present else amount_col
    filter_col = due_col if due_col else amount_col
    if filter_col:
        # We don't coerce display strings; keep original raw values to decide
        # Build due_invoices: where parsed numeric > 0
        parsed_list = []
        for v in client_invoices[filter_col].tolist():
            c, a = parse_currency(v)
            parsed_list.append(a)
        client_invoices["_parsed_num"] = parsed_list
        due_invoices = client_invoices[client_invoices["_parsed_num"] > 0].copy()
    else:
        due_invoices = client_invoices

    num_due = len(due_invoices)

# Determine primary currency for client (from first invoice)
first_currency = "INR"
if len(due_invoices) > 0:
    first_currency = due_invoices["Currency"].iloc[0] if "Currency" in due_invoices.columns else "INR"

client_total = 0.0
invoice_rows = ""

for idx, row in due_invoices.iterrows():
    invoice_num = row[invoice_col] if invoice_col else "-"
    currency_val = row.get("Currency", first_currency)
    raw_val = row[amount_col] if amount_col in row.index else row.get(amount_col, 0)

    # Parse numeric amount
    _, amt = parse_currency(raw_val)

    # Convert to primary currency if needed
    if currency_val != first_currency:
        if first_currency == "INR":
            amt = convert_to_inr(currency_val, amt)
        else:
            try:
                amt = float(amt) / st.session_state.USD_TO_INR
            except Exception:
                amt = 0.0

    client_total += amt
    amount_val = f"{amt:,.2f}"  # Display numeric only, no symbol

    invoice_date_val = (
        row[date_col].strftime("%Y-%m-%d") if pd.notnull(row[date_col]) else "-"
    )
    days_pending = (
        (datetime.now() - pd.to_datetime(row[date_col], errors="coerce")).days
        if pd.notnull(row[date_col])
        else "-"
    )

    # Build table row with Currency column
    invoice_rows += f"""
    <tr>
        <td>{invoice_num}</td>
        <td>{currency_val}</td>
        <td>{amount_val}</td>
        <td>{invoice_date_val}</td>
        <td>{days_pending} days</td>
    </tr>
    """

# Total line with client's primary currency
total_line = f"<li><strong>Total Amount Due ({first_currency}):</strong> {client_total:,.2f}</li>"

# Updated email table header with Currency column
auto_message = f"""
<html>
<body style="background-color: none; color: #ffffff;">
<p>Dear Sir/Mam,</p>
<p>Please find below your pending invoices:</p>

<table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse; width: 100%;">
    <thead style="background-color: none;">
        <tr>
            <th>Invoice #</th>
            <th>Currency</th>
            <th>Amount</th>
            <th>Invoice Date</th>
            <th>Days Pending</th>
        </tr>
    </thead>
    <tbody>{invoice_rows}</tbody>
</table>

<ul style="margin-top: 10px;">
    <li><strong>No. of Due Invoices:</strong> {num_due}</li>
    {total_line}
</ul>

<p>Kindly arrange the payments at the earliest convenience.</p>
<p>Thanks & Regards,<br>S2 Integrators</p>
</body>
</html>
"""

    # Dynamic subject for each client
email_subject = f"{selected_client_name} - Pending Invoice Payment | S2 Integrators Pvt Ltd"

    # 🟢 Reset the text fields each time a different client is selected
st.session_state.email_subject = email_subject
st.session_state.email_message = auto_message

subject_input = st.sidebar.text_input("Email Subject", value=st.session_state.email_subject, key=f"email_subject_{selected_client_name}")

# --- Email Message Section ---
st.sidebar.markdown("## 💬 Email Message")

with st.sidebar.expander("✏️ Edit Email (HTML Code)", expanded=False):
    try:
        _ = selected_client_name  # noqa: F841
    except NameError:
        selected_client_name = ""
    message_input = st.text_area("Email Message (HTML)", value=st.session_state.get("email_message", ""), height=300, key=f"email_message_{selected_client_name}")

import streamlit.components.v1 as components

with st.sidebar.expander("👁️ Preview Formatted Email", expanded=False):
    components.html(message_input, height=400, scrolling=True)

# Resolve client and cc email using stored config first, then Excel fallback
try:
    client_name = selected_client_name
except NameError:
    client_name = ""

client_email = None
cc_email = None

if "client_emails" in st.session_state:
    client_email = st.session_state.client_emails.get("clients", {}).get(client_name, None)
    cc_email = st.session_state.client_emails.get("cc_email", None)

if not client_email and 'client_mail_col' in locals() and client_mail_col:
    try:
        client_email = client_invoices[client_mail_col].iloc[0]
    except Exception:
        client_email = None
if not cc_email and 'cc_mail_col' in locals() and cc_mail_col:
    try:
        cc_email = client_invoices[cc_mail_col].iloc[0]
    except Exception:
        cc_email = None

# --- Send button ---
st.sidebar.markdown("## ᯓ➤ Send Mail to Client")

if num_due == 0:
    st.sidebar.warning("✅ No pending invoices for this client. Email not required.")
else:
    if st.sidebar.button("🚀 Send Now"):
        if not st.session_state.sender_email or not st.session_state.sender_password:
            st.sidebar.warning("⚠️ Please set sender credentials!")
        elif client_email:
            def send_email_html(sender_email, sender_password, to_email, subject, html_body, cc=None):
                msg = EmailMessage()
                msg["Subject"] = subject
                msg["From"] = sender_email
                msg["To"] = to_email
                if cc:
                    msg["Cc"] = cc
                msg.add_alternative(html_body, subtype="html")
                try:
                    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
                        smtp.login(sender_email, sender_password)
                        smtp.send_message(msg)
                    return True, "Email sent successfully!"
                except Exception as e:
                    return False, str(e)

            success, msg = send_email_html(st.session_state.sender_email, st.session_state.sender_password, client_email, subject_input, message_input, cc=cc_email)

            if success:
                st.sidebar.success(f"✅ Email sent to {client_email}")
                st.toast(f"Email sent to {client_email}", icon="✅")
            else:
                st.sidebar.error(f"❌ Failed to send email: {msg}")
                st.toast(f"Failed to send email: {msg}", icon="❌")
        else:
            st.sidebar.error("⚠️ Client email address not found.")
            st.toast(f"Client email address not found.", icon="⚠️")
